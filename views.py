from datetime import datetime
from io import BytesIO
from unittest import case
from django.shortcuts import render, redirect, get_object_or_404
from django.db import models
from django import forms
from django.core.paginator import Paginator
from django.views import View
from portal.models import BaseDados, Produto, Precos_graduacao_vinculo, Precos_status_graduacao
from .forms import ReservasForm
from django.http import HttpResponse, HttpResponseBadRequest, HttpResponseNotAllowed, JsonResponse
from django.http import HttpResponseRedirect
from django.urls import reverse
from django.views.decorators.csrf import csrf_protect
import logging
from decimal import Decimal
from django.template.defaulttags import register
from django.db.models import Sum
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
import locale
from datetime import datetime
from django.db.models import F
from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Side, Alignment
import xlsxwriter
from django.http import Http404
from django.db.models import Case, When, Value, F, IntegerField, Q
from django.views.generic import View
import json
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required

@register.filter
def get_range(value):
    return range(value)


@login_required
def consultar_reservas(request):
    consultar = BaseDados.objects.filter(status_reserva="Pendente")  # Filtra por status "pendente"
    print(consultar)    
    context = {
        'consultar': consultar,
    }
    return render(request, 'portal/reservas.html', context)

def reserva_externa(request):
    if request.method == 'POST':
        form = ReservasForm(request.POST)
        if form.is_valid():
            obj = form.save()
            print("Objeto salvo:", obj)
            form.save()
            # return redirect('consultar_reservas')
            return redirect('reserva_externa')
        else:
            # Se o formulário for inválido, cairemos aqui e os erros serão impressos.
            print("Formulário inválido", form.errors)
    else:
        # Se a requisição não for POST, um formulário em branco será fornecido.
        form = ReservasForm()

    # O contexto é passado para o template tanto se o formulário for válido (e redirecionado antes),
    # quanto se não for POST ou se for inválido.
    context = {'form': form}
    return render(request, 'portal/index.html', context)

@login_required
def fazer_reservas(request):
    if request.method == 'POST':
        form = ReservasForm(request.POST)
        if form.is_valid():
            obj = form.save()
            print("Objeto salvo:", obj)
            form.save()
            # return redirect('consultar_reservas')
            return redirect('consultar_reservas')
        else:
            # Se o formulário for inválido, cairemos aqui e os erros serão impressos.
            print("Formulário inválido", form.errors)
    else:
        # Se a requisição não for POST, um formulário em branco será fornecido.
        form = ReservasForm()

    # O contexto é passado para o template tanto se o formulário for válido (e redirecionado antes),
    # quanto se não for POST ou se for inválido.
    context = {'form': form}
    return render(request, 'portal/reserva_add.html', context)

@login_required
def editar_reservas(request, reservas_pk):
    # Aqui usamos get_object_or_404 para garantir que se o objeto não for encontrado,
    # uma página de erro 404 será retornada.
    editar = get_object_or_404(BaseDados, pk=reservas_pk)
    
    # Se estamos lidando com uma requisição POST, significa que o formulário foi submetido
    if request.method == 'POST':
        # Passamos a instância da reserva que queremos editar para o formulário,
        # junto com os dados submetidos (request.POST)
        form = ReservasForm(request.POST, instance=editar)
        print("Dados recebidos do formulário:", request.POST)
        if 'status_reserva' in request.POST:
            status_reserva = request.POST['status_reserva']
            print("Status da reserva recebido:", status_reserva)
            # Atribuir o valor do campo status_reserva à instância do formulário
            form.instance.status_reserva = status_reserva
        
        # Verificamos se o formulário é válido
        if form.is_valid():
            # Se for válido, salvamos as alterações feitas na reserva
            form.save()
            # E redirecionamos o usuário para a página de consulta das reservas
            return redirect('consultar_reservas')
    else:
        # Se a requisição for um GET, criamos um formulário preenchido com as informações
        # da reserva que queremos editar
        form = ReservasForm(instance=editar)
    
    # Passamos o formulário preenchido para o template
    context = {'form': form, 'reserva': editar}
    return render(request, 'portal/reserva_edit.html', context)

@login_required
def recepcao(request):
    consultar = BaseDados.objects.filter(status_reserva="Aprovada")  # Filtra por status "pendente"
    consultar_checkin = BaseDados.objects.filter(status_reserva="Checkin")
    # print(consultar)    
    context = {
        'consultar': consultar,
        'consultar_checkin': consultar_checkin,
    }
    return render(request, 'portal/recepcao.html', context)    



@csrf_protect
@login_required
def editar_checkin(request, reservas_pk):
    reserva = get_object_or_404(BaseDados, pk=reservas_pk)
    
    if request.method == 'POST':
        novo_status = request.POST.get('status_reserva')
        reserva.status_reserva = novo_status
        
        # Calcula as diárias
        if reserva.entrada and reserva.saida:
            try:
                formato_data = "%d/%m/%Y"
                data_entrada_str = reserva.entrada.strftime(formato_data)  # Convertendo para string
                data_saida_str = reserva.saida.strftime(formato_data)  # Convertendo para string

                data_entrada = datetime.strptime(data_entrada_str, formato_data)
                data_saida = datetime.strptime(data_saida_str, formato_data)
                
                diarias = (data_saida - data_entrada).days
                reserva.diarias = diarias
                print(data_entrada)
                print(data_saida)
                print(reserva.diarias)
            except ValueError as e:
                print(f"Erro ao calcular diárias: {e}")
                reserva.diarias = 0
        
        reserva.save()
        
        return HttpResponseRedirect(reverse('recepcao'))
    
    return render(request, 'portal/checkin.html', {'reserva': reserva})



def editar_checkout(request, reservas_pk):
    reserva = get_object_or_404(BaseDados, pk=reservas_pk)
    
    if request.method == 'POST':
        novo_status = request.POST.get('status_reserva')
        reserva.status_reserva = novo_status
        
        # Verifica se o checkbox foi marcado
        if 'pagante_checkbox' in request.POST:
            print("Checkbox foi marcado")
            nome_pagante = request.POST.get('nome_pagante')
            cpf_pagante = request.POST.get('cpf_pagante')
            print("Nome do pagante:", nome_pagante)
            print("CPF do pagante:", cpf_pagante)
        else:
            # Se o checkbox não foi marcado, usa os dados existentes
            print("Checkbox não foi marcado")
            nome_pagante = reserva.nome
            cpf_pagante = reserva.cpf
        
        # Atualiza os campos na reserva
        reserva.nome_pagante = nome_pagante
        reserva.cpf_pagante = cpf_pagante
        
        # Salva a reserva
        reserva.save()
        
        return HttpResponseRedirect(reverse('recepcao'))
    
    return render(request, 'portal/checkout.html', {'reserva': reserva})



@csrf_protect

def editar_consumacao(request, reservas_pk):
    reserva = get_object_or_404(BaseDados, pk=reservas_pk)

    print("Valores do objeto reserva antes de renderizar o formulário:", reserva.qtde_agua, reserva.qtde_refri, reserva.qtde_cerveja, reserva.status, reserva.graduacao, reserva.valor_hosp)
    
    if request.method == 'POST':
        novo_agua = request.POST.get('qtde_agua')
        novo_refri = request.POST.get('qtde_refri')
        novo_cerveja = request.POST.get('qtde_cerveja')

        status = reserva.status
        graduacao = reserva.graduacao

        # Consulta os preços com base no status e na graduação
        try:
            if status == 'CIVIL':
                preco_status_graduacao = Precos_status_graduacao.objects.get(status=status, graduacao=graduacao)
                reserva.valor_hosp = preco_status_graduacao.valor
            elif status in ['MILITAR DA ATIVA', 'MILITAR DA RESERVA', 'DEP ACOMPANHADO', 'DEP DESACOMPANHADO', 'PENSIONISTA']:                
                preco_status_graduacao = Precos_status_graduacao.objects.get(status=status, graduacao=graduacao)
                reserva.valor_hosp = preco_status_graduacao.valor
            else:
                # Se não encontrar um preço correspondente, define valor_hosp como zero
                reserva.valor_hosp = 0  # Ou algum outro valor padrão que faça sentido em seu contexto
        except Precos_status_graduacao.DoesNotExist:
            # Se o objeto não for encontrado, define valor_hosp como zero
            reserva.valor_hosp = 0
        
        # Exemplo de cálculo para um acompanhante
        vinculo_acomps = [reserva.vinculo_acomp1, reserva.vinculo_acomp2, reserva.vinculo_acomp3, reserva.vinculo_acomp4, reserva.vinculo_acomp5]
        for i, vinculo_acomp in enumerate(vinculo_acomps):
            if vinculo_acomp:
                try:
                    # Consulta o preço do vínculo e graduação na tabela Precos_graduacao_vinculo
                    preco_vinculo = Precos_graduacao_vinculo.objects.get(graduacao=graduacao, vinculo=vinculo_acomp)
                    setattr(reserva, f'valor_acomp{i+1}', preco_vinculo.valor)
                except Precos_graduacao_vinculo.DoesNotExist:
                    setattr(reserva, f'valor_acomp{i+1}', 0)
        # Cálculo dos novos valores conforme solicitado
        if isinstance(reserva.valor_hosp, (int, float, Decimal)):
            reserva.valor_dia = (reserva.valor_hosp + reserva.valor_acomp1 + reserva.valor_acomp2 + reserva.valor_acomp3 + reserva.valor_acomp4 + reserva.valor_acomp5)
        else:
            # Se reserva.valor_hosp não for numérico, defina reserva.valor_dia como zero ou algum outro valor padrão
            reserva.valor_dia = 0
                   

        # print("Tipo de dado do valor:", type(preco_status_graduacao.valor))
        print("Valor do hospede:", reserva.valor_hosp)
        print("Valores recebidos do formulário:")
        print("Água:", novo_agua)
        print("Refri:", novo_refri)
        print("Cerveja:", novo_cerveja)
        reserva.qtde_acomp = reserva.qtde_hosp - 1
        
        # Verifica se os valores não são None antes de converter para Decimal
        if novo_agua is not None:
            reserva.qtde_agua = Decimal(novo_agua)
        if novo_refri is not None:
            reserva.qtde_refri = Decimal(novo_refri)
        if novo_cerveja is not None:
            reserva.qtde_cerveja = Decimal(novo_cerveja)

        
            
        # Busca os valores dos produtos no banco de dados
        produto_agua = Produto.objects.get(nome='Água')
        produto_refri = Produto.objects.get(nome='Refrigerante')
        produto_cerveja = Produto.objects.get(nome='Cerveja')   
        
        if reserva.motivo_viagem == "Saúde":
            reserva.desc_saude = reserva.valor_dia / 2
            reserva.subtotal = reserva.desc_saude * reserva.diarias          
            reserva.valor_total = reserva.subtotal + reserva.valor_ajuste + reserva.total_consumacao  
            reserva.total_agua = produto_agua.valor * reserva.qtde_agua if reserva.qtde_agua is not None else Decimal(0)
            reserva.total_refri = produto_refri.valor * reserva.qtde_refri if reserva.qtde_refri is not None else Decimal(0)
            reserva.total_cerveja = produto_cerveja.valor * reserva.qtde_cerveja if reserva.qtde_cerveja is not None else Decimal(0)
            reserva.total_consumacao = reserva.total_agua + reserva.total_refri + reserva.total_cerveja
            reserva.valor_total = reserva.subtotal + reserva.valor_ajuste + reserva.total_consumacao
        else:
            reserva.desc_saude = 0
            reserva.subtotal = reserva.valor_dia * reserva.diarias          
            reserva.valor_total = reserva.subtotal + reserva.valor_ajuste + reserva.total_consumacao  
            reserva.total_agua = produto_agua.valor * reserva.qtde_agua if reserva.qtde_agua is not None else Decimal(0)
            reserva.total_refri = produto_refri.valor * reserva.qtde_refri if reserva.qtde_refri is not None else Decimal(0)
            reserva.total_cerveja = produto_cerveja.valor * reserva.qtde_cerveja if reserva.qtde_cerveja is not None else Decimal(0)
            reserva.total_consumacao = reserva.total_agua + reserva.total_refri + reserva.total_cerveja
            reserva.valor_total = reserva.subtotal + reserva.valor_ajuste + reserva.total_consumacao    


        # Calcula consumacao
        reserva.total_agua = produto_agua.valor * reserva.qtde_agua if reserva.qtde_agua is not None else Decimal(0)
        reserva.total_refri = produto_refri.valor * reserva.qtde_refri if reserva.qtde_refri is not None else Decimal(0)
        reserva.total_cerveja = produto_cerveja.valor * reserva.qtde_cerveja if reserva.qtde_cerveja is not None else Decimal(0)
        reserva.total_consumacao = reserva.total_agua + reserva.total_refri + reserva.total_cerveja
        reserva.valor_total = reserva.subtotal + reserva.valor_ajuste + reserva.total_consumacao
        
        reserva.save()
        
        return HttpResponseRedirect(reverse('recepcao'))
    
    return render(request, 'portal/consumacao.html', {'reserva': reserva})




@login_required
def relatorio_mensal(request):
    if request.method == 'GET':
        mes_relatorio = request.GET.get('mes_relatorio')
        ano_relatorio = request.GET.get('ano_relatorio')
        forma_pagamento = request.GET.get('forma_pagamento')
        print("Mês Relatório:", mes_relatorio)  # Verifique se os valores estão corretos
        print("Ano Relatório:", ano_relatorio)
        print("Forma Pagamento:", forma_pagamento)
        
        # Verificar se mes_relatorio e ano_relatorio não são None
        if mes_relatorio is None or ano_relatorio is None:
            return HttpResponseBadRequest("Parâmetros mes_relatorio e ano_relatorio são obrigatórios.")

        # Validação dos dados de entrada
        try:
            mes_relatorio = int(mes_relatorio)
            ano_relatorio = int(ano_relatorio)
        except ValueError:
            return HttpResponseBadRequest("Mês e ano devem ser números inteiros.")
        
        if mes_relatorio < 1 or mes_relatorio > 12:
            return HttpResponseBadRequest("Mês deve estar entre 1 e 12.")
        
        if ano_relatorio < 2024 or ano_relatorio > 2030:
            return HttpResponseBadRequest("Ano deve estar entre 2024 e 2030.")

        # Realize as consultas usando os parâmetros recebidos
        consultar_htm1 = BaseDados.objects.filter(
            status_reserva="Pago",
            mhex="HTM_01",
            forma_pagamento=forma_pagamento,
            saida__month=mes_relatorio,
            saida__year=ano_relatorio
        )

        total_htm1 = BaseDados.objects.filter(
            status_reserva="Pago",
            mhex="HTM_01",
            forma_pagamento=forma_pagamento,
            saida__month=mes_relatorio,
            saida__year=ano_relatorio
        ).aggregate(soma=Sum('valor_total'))['soma'] 

          

        consultar_htm2 = BaseDados.objects.filter(
            status_reserva="Pago",
            mhex="HTM_02",
            forma_pagamento=forma_pagamento,
            saida__month=mes_relatorio,
            saida__year=ano_relatorio
        )

        total_htm2 = BaseDados.objects.filter(
            status_reserva="Pago",
            mhex="HTM_02",
            forma_pagamento=forma_pagamento,
            saida__month=mes_relatorio,
            saida__year=ano_relatorio
        ).aggregate(soma=Sum('valor_total'))['soma'] 

        # Atualização em lote dos objetos BaseDados
        if consultar_htm1.exists():
            consultar_htm1.update(qtde_acomp=Case(
                When(qtde_hosp=1, then=Value(0)),
                default=F('qtde_hosp') - 1,
                output_field=IntegerField()
            ))

        if consultar_htm2.exists():
            consultar_htm2.update(qtde_acomp=Case(
                When(qtde_hosp=1, then=Value(0)),
                default=F('qtde_hosp') - 1,
                output_field=IntegerField()
            ))

        context = {
            'consultar_htm1': consultar_htm1,
            'consultar_htm2': consultar_htm2,
            'total_htm1': total_htm1,
            'total_htm2': total_htm2,

        }
        return render(request, 'portal/relatorio_mensal.html', context)
    else:
        return HttpResponseBadRequest("Método de requisição inválido.")

@login_required
def relatorio_pagamento(request):
    consultar = BaseDados.objects.filter(status_reserva="Pago")
    
    context = {
        'consultar': consultar,
        
    }
    return render(request, 'portal/relatorio_pagamento.html', context) 

@login_required
def relatorio_pagamento_pix(request):
    consultar = BaseDados.objects.filter(status_reserva="Pago")
    
    context = {
        'consultar': consultar,
        
    }
    return render(request, 'portal/relatorio_pagamento_pix.html', context) 

@login_required
def relatorio_pagamento_dinheiro(request):
    consultar = BaseDados.objects.filter(status_reserva="Pago")
    
    context = {
        'consultar': consultar,
        
    }
    return render(request, 'portal/relatorio_pagamento_dinheiro.html', context) 





@login_required
def relatorio_pagamento_excel(request):
    consultar = BaseDados.objects.filter(status_reserva="Pago")
    
    context = {
        'consultar': consultar,
        
    }
    return render(request, 'portal/relatorio_pagamento_excel.html', context) 



class RelatorioPix(View):
    nome_colunas_excel = {
        'saida': 'Data de Saída',
        'graduacao': 'PST/GRAD',
        'nome_pagante': 'Hóspede',
        'cpf_pagante': 'CPF',
        'qtde_acomp': 'Nr Acomp',
        'uh': 'UH',
        'diarias': 'Dias',
        'valor_total': 'Total (R$)',
    }

    cabecalho_fixo = 9
    cabecalho_dados1 = 1
    cabecalho_dados2 = 1
    linhas_dados1 = None
    linhas_dados2 = None
    mes_texto = None
    total_linhas = None
    total_htm1 = 0
    total_htm2 = 0
    dados1 = None
    dados2 = None
    mes_relatorio = None
    ano_relatorio = None


    def get(self, request, *args, **kwargs):
        self.mes_relatorio = request.GET.get('mes_relatorio')
        self.ano_relatorio = request.GET.get('ano_relatorio')
        forma_pagamento = request.GET.get('forma_pagamento')
        # print("Mês Relatório: auditoria", mes_relatorio)  # Verifique se os valores estão corretos
        # print("Ano Relatório: auditoria", ano_relatorio)
        # print("Forma Pagamento: auditoria", forma_pagamento)

        meses = {
            '1': 'JAN',
            '2': 'FEV',
            '3': 'MAR',
            '4': 'ABR',
            '5': 'MAI',
            '6': 'JUN',
            '7': 'JUL',
            '8': 'AGO',
            '9': 'SET',
            '10': 'OUT',
            '11': 'NOV',
            '12': 'DEZ'
        }

        
        # Verificar se mes_relatorio, ano_relatorio e forma_pagamento não são None
        if self.mes_relatorio is None or self.ano_relatorio is None or forma_pagamento is None:
            return HttpResponseBadRequest("Parâmetros mes_relatorio, ano_relatorio e forma_pagamento são obrigatórios.")

        # Validação dos dados de entrada
        try:
            self.mes_relatorio = int(self.mes_relatorio)
            self.ano_relatorio = int(self.ano_relatorio)
        except ValueError:
            return HttpResponseBadRequest("Mês e ano devem ser números inteiros.")

        if self.mes_relatorio < 1 or self.mes_relatorio > 12:
            return HttpResponseBadRequest("Mês deve estar entre 1 e 12.")

        if self.ano_relatorio < 2024 or self.ano_relatorio > 2030:
            return HttpResponseBadRequest("Ano deve estar entre 2024 e 2030.")
        
        self.mes_texto = meses.get(str(self.mes_relatorio), 'DESC')
        print(self.mes_texto)

        # Realize as consultas usando os parâmetros recebidos
        consultar_htm1 = BaseDados.objects.filter(
            status_reserva="Pago",
            mhex="HTM_01",
            forma_pagamento=forma_pagamento,
            saida__month=self.mes_relatorio,
            saida__year=self.ano_relatorio
        )

        # Consulta para o segundo tipo de dados
        consultar_htm2 = BaseDados.objects.filter(
            status_reserva="Pago",
            mhex="HTM_02",
            forma_pagamento=forma_pagamento,
            saida__month=self.mes_relatorio,
            saida__year=self.ano_relatorio
        )

        if not consultar_htm1.exists() and not consultar_htm2.exists():
            return render(request, 'portal/relatorio_pagamento_pix.html', {'exibir_modal': True})

        self.total_htm1 = BaseDados.objects.filter(
            status_reserva="Pago",
            mhex="HTM_01",
            forma_pagamento=forma_pagamento,
            saida__month=self.mes_relatorio,
            saida__year=self.ano_relatorio
        ).aggregate(soma=Sum('valor_total'))['soma']  

        self.total_htm2 = BaseDados.objects.filter(
            status_reserva="Pago",
            mhex="HTM_02",
            forma_pagamento=forma_pagamento,
            saida__month=self.mes_relatorio,
            saida__year=self.ano_relatorio
        ).aggregate(soma=Sum('valor_total'))['soma']

        # Combine os resultados das consultas em um único conjunto de dados
        self.dados1 = list(consultar_htm1)
        self.dados2 = list(consultar_htm2)

        self.linhas_dados1 = len(self.dados1) 
        self.linhas_dados2 = len(self.dados2)       
        self.total_linhas = self.linhas_dados1 + self.linhas_dados2

        # Gere a planilha
        workbook = self.gerar_planilha(self.dados1, self.dados2, self.linhas_dados1, self.mes_relatorio, self.mes_texto)

        # Configura o Excel
        self.configurar_excel(workbook, self.dados1, self.dados2, self.linhas_dados1, self.linhas_dados2, self.total_linhas, self.ano_relatorio)

        # Configura o Estilo
        self.configurar_estilos(workbook, self.dados1, self.dados2, self.linhas_dados1, self.linhas_dados2, self.total_linhas, self.total_htm1, self.total_htm2, self.ano_relatorio)

        # Retorna a resposta HTTP com a planilha anexada
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="relatorio_{forma_pagamento}.xlsx"'

        # Salva o workbook como uma resposta HTTP
        workbook.save(response)
        return response

    def gerar_planilha(self, dados1, dados2, linhas_dados1, mes_relatorio, mes_texto):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = f'{ self.mes_texto}'

        # INCLUIR CABEÇALHO DADOS 1
        for col, coluna in enumerate(self.nome_colunas_excel.values(), start=1):
            sheet.cell(row=10, column=col).value = coluna

        # INCLUIR DADOS 1
        self.escrever_dados(sheet, 11, self.dados1)

        # INCLUIR CABEÇALHO DADOS 2
        for col, coluna in enumerate(self.nome_colunas_excel.values(), start=1):
            sheet.cell(row=self.linhas_dados1 + 10 + 3, column=col).value = coluna

        # Escreva os dados do segundo conjunto na planilha a partir da linha len(dados1) + 11 + 3
        self.escrever_dados(sheet, self.linhas_dados1 + 10 + 4, self.dados2)

        return workbook

    def escrever_dados(self, sheet, start_row, dados):
        # Escrever os dados
        for row, registro in enumerate(dados, start=start_row):
            for col, coluna in enumerate(self.nome_colunas_excel.keys(), start=1):
                valor = getattr(registro, coluna, None)
                if coluna == 'valor_total' and valor is not None:
                    try:
                        valor = float(valor)
                    except ValueError:
                        valor = None
                sheet.cell(row=row, column=col).value = valor

    def configurar_excel(self, workbook, dados1, dados2, linhas_dados1, linhas_dados2, total_linhas, ano_relatorio):
        # Acessa a planilha ativa
        sheet = workbook.active

        # Larguras das colunas desejadas
        larguras_colunas = {'A': 19, 'B': 13, 'C': 56, 'D': 16.71, 'E': 10, 'F': 9.5, 'G': 9.5, 'H': 17}
        for col, width in larguras_colunas.items():
            sheet.column_dimensions[col].width = width

        # Formatar coluna 'A' (Saida) para dd/mm/aaaa
        date_format = NamedStyle(name='date_format')
        date_format.number_format = 'DD/MM/YYYY'

        # Aplicar estilo à coluna 'A' (Saida)
        for cell in sheet['A']:
            cell.style = date_format

        # Ocultar as linhas de grade
        sheet.sheet_view.showGridLines = False

        # Formatar coluna 'H' (valor_total) para ##.###,##
        valor_format = NamedStyle(name='valor_format')
        valor_format.number_format = '#,##0.00' if '.' in locale.localeconv()['decimal_point'] else '#.##0,00'

        # Aplicar estilo à coluna 'H' (valor_total)
        for cell in sheet['H']:
            cell.style = valor_format

        # Consulta para os dados do banco
        consultar_dados = self.dados1

        # Configura outros estilos, como bordas, fontes, etc.
        self.configurar_estilos(workbook, self.dados1, self.dados2, self.linhas_dados1, self.linhas_dados2, self.total_linhas, self.total_htm1, self.total_htm2, self.ano_relatorio)
        
        # Centralizar as colunas especificadas
        colunas_centralizadas = ['A', 'B', 'D', 'E', 'F', 'G']  # Colunas 'saida', 'graduacao', 'cpf', 'qtde_acomp', 'uh', 'diarias'
        for coluna in colunas_centralizadas:
            for cell in sheet[coluna]:
                cell.alignment = Alignment(horizontal='center')

        # consultar_htm1 = BaseDados.objects.filter(status_reserva="Pago", mhex="HTM_01", forma_pagamento="PIX", saida__month=1)
        # dados1 = list(consultar_htm1)
        # consultar_htm2 = BaseDados.objects.filter(status_reserva="Pago", mhex="HTM_02", forma_pagamento="PIX", saida__month=1)
        # dados2 = list(consultar_htm2)
        # linhas_dados1 = len(dados1) 
        # linhas_dados2 = len(dados2)       
         
        sheet.cell(row=self.total_linhas + 15, column=1).alignment = Alignment(horizontal='left', vertical='center')
        sheet.cell(row=self.total_linhas + 17, column=1).alignment = Alignment(horizontal='left', vertical='center')
        sheet.cell(row=self.total_linhas + 18, column=1).alignment = Alignment(horizontal='left', vertical='center')        

    # def escrever_dados(self, sheet, start_row, dados):
    #     # Defina as colunas desejadas diretamente aqui
    #     colunas_desejadas = ['saida', 'graduacao', 'nome', 'cpf', 'qtde_acomp', 'uh', 'diarias', 'valor_total']

    #     # Escrever os dados
    #     for row, registro in enumerate(dados, start=start_row):
    #         for col, coluna in enumerate(colunas_desejadas, start=1):
    #             valor = getattr(registro, coluna, None)
    #             if coluna == 'valor_total' and valor is not None:
    #                 try:
    #                     valor = float(valor)
    #                 except ValueError:
    #                     valor = None
    #             sheet.cell(row=row, column=col).value = valor

    def configurar_estilos(self, workbook, dados1, dados2, linhas_dados1, linhas_dados2, total_linhas, total_htm1, total_htm2, ano_relatorio):
        sheet = workbook.active
        # consultar_htm1 = BaseDados.objects.filter(status_reserva="Pago", mhex="HTM_01", forma_pagamento="PIX", saida__month=1)
        # dados1 = list(consultar_htm1)
        # consultar_htm2 = BaseDados.objects.filter(status_reserva="Pago", mhex="HTM_02", forma_pagamento="PIX", saida__month=1)
        # dados2 = list(consultar_htm2)
        
        # linhas_dados1 = len(dados1) 
        # linhas_dados2 = len(dados2)       
        # total_linhas = linhas_dados1 + linhas_dados2

        
        # total_linhas = len(dados1) + len(dados2)
        sheet.cell(row=self.total_linhas + 15, column=1).alignment = Alignment(horizontal='left', vertical='center')
        
        meses2 = {
            'JAN': 'janeiro',
            'FEV': 'fevereiro',
            'MAR': 'março',
            'ABR': 'abril',
            'MAI': 'maio',
            'JUN': 'junho',
            'JUL': 'julho',
            'AGO': 'agosto',
            'SET': 'setembro',
            'OUT': 'outubro',
            'NOV': 'novembro',
            'DEZ': 'dezembro'
        }

        self.mes2_texto = meses2.get(self.mes_texto)

        # Defina os estilos de borda
        dotted_border = Border(left=Side(style='dotted'),
                            right=Side(style='dotted'),
                            top=Side(style='dotted'),
                            bottom=Side(style='dotted'))

        thin_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))

        # Define as fontes
        fonte_padrao = Font(name='Times New Roman', size=11, bold=False)
        fonte_negrito = Font(name='Times New Roman', size=11, bold=True)

        # Define o preenchimento cinza claro para o cabeçalho
        gray_fill = PatternFill(start_color='00CCCCCC',
                                end_color='00CCCCCC',
                                fill_type='solid')

        # Aplica os estilos para todas as células
        for row in sheet.iter_rows():
            for cell in row:
                # Aplica a borda
                cell.border = thin_border

                # Aplica a fonte padrão
                cell.font = fonte_padrao

        # Aplica estilos para o cabeçalho
        for row in sheet.iter_rows(min_row=10, max_row=10):
            for cell in row:
                cell.font = fonte_negrito
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = gray_fill

        # Aplica estilos para o cabeçalho 2
        for row in sheet.iter_rows(min_row=len(dados1) + 11 + 2, max_row=len(dados1) + 11 + 2):
            for cell in row:
                cell.font = fonte_negrito
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = gray_fill

        # # Defina o estilo de borda externa
        # outer_border = Border(left=Side(style='thin'), right=Side(style='thin'),
        #                     top=Side(style='thin'), bottom=Side(style='thin'))
        # # Aplica a borda externa às células das colunas 1 a 8 na linha total_linhas + 18
        # for col in range(1, 9):
        #     sheet.cell(row=self.total_linhas + 13, column=col).border = outer_border
        #     sheet.cell(row=self.total_linhas + 14, column=col).border = outer_border 
        #     sheet.cell(row=self.total_linhas + 15, column=col).border = outer_border 
        #     sheet.cell(row=self.total_linhas + 16, column=col).border = outer_border 
        #     sheet.cell(row=self.total_linhas + 17, column=col).border = outer_border 
        #     sheet.cell(row=self.total_linhas + 18, column=col).border = outer_border 
            
        # Seu texto
        texto = """MINISTERIO DA DEFESA
EXÉRCITO BRASILEIRO
COMANDO MILITAR DO OESTE
4ª BRIGADA DE CAVALARIA MECANIZADA
11º REGIMENTO DE CAVALARIA MECANIZADO
REGIMENTO MARECHAL DUTRA
"""

        # Insere o texto na célula A1
        sheet['A1'] = """MINISTERIO DA DEFESA"""
        sheet['A2'] = """EXÉRCITO BRASILEIRO"""   
        sheet['A3'] = """COMANDO MILITAR DO OESTE"""   
        sheet['A4'] = """4ª BRIGADA DE CAVALARIA MECANIZADA"""   
        sheet['A5'] = """11º REGIMENTO DE CAVALARIA MECANIZADO"""   
        sheet['A6'] = """REGIMENTO MARECHAL DUTRA"""   
        sheet['A7'] = """RELATÓRIO DE AUDITORIA DOS MEIOS DE HOSPEDAGEM DO EXÉRCITO NA GUARNIÇÃO DE PONTA PORÃ - MS"""   
        sheet['A8'] = f"Trata o presente relatório sobre auditoria realizada nos meios de Hospedagem do Exército da Guarnição de Ponta Porã relativo ao mês de {self.mes2_texto} de {self.ano_relatorio}."  
        sheet['A9'] = """HTM 1"""
        sheet.cell(row=self.linhas_dados1 + 11, column=8).value = total_htm1
        sheet.cell(row=self.linhas_dados1 + 12, column=1).value = "HTM 2"
        
        
        
        
       
        sheet.cell(row=total_linhas + 14, column=8).value = total_htm2
        
        total_geral = total_htm1 + total_htm2
        sheet.cell(row=total_linhas + 15, column=1).value = "TOTAL EM PIX"
        sheet.cell(row=total_linhas + 15, column=8).value = total_geral

        sheet.cell(row=total_linhas + 17, column=1).value = "2. RECOLHIDO AO FUNDO DO EXÉRCITO"
        sheet.cell(row=total_linhas + 18, column=1).value = "DEP. PIX"
        sheet.cell(row=total_linhas + 18, column=8).value = total_geral

        # Define o idioma como português brasileiro
        locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
        # Obter a data de hoje
        data_hoje = datetime.now()
        # Formatar a data como "dia de mês de ano"
        data_formatada = data_hoje.strftime("%d de %B de %Y")

        # Inserir a data formatada na célula desejada
        sheet.cell(row=total_linhas + 20, column=1).value = f"Ponta Porã-MS, {data_formatada}"
        sheet.cell(row=total_linhas + 23, column=1).value = "EDER SCHWEIGERT FONSECA - Cap"
        sheet.cell(row=total_linhas + 24, column=1).value = "Gestor HT 11º R C Mec"
        sheet.cell(row=total_linhas + 27, column=1).value = "CHRYSTIAN HENRY BRITO CARDOSO - Maj"
        sheet.cell(row=total_linhas + 28, column=1).value = "Fiscal Administrativo 11º R C Mec"

        

        

        # Ajusta o alinhamento para envolver o texto e centralizá-lo
        sheet['A1'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A2'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A3'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A4'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A5'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A6'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A7'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A8'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A9'].alignment = Alignment(wrapText=False, horizontal='center', vertical='center')
        sheet.cell(row=len(dados1) + 12, column=1).alignment = Alignment(wrapText=False, horizontal='center', vertical='center')

        

        # Mescla as células A1:H1
        sheet.merge_cells('A1:H1')
        sheet.merge_cells('A2:H2')
        sheet.merge_cells('A3:H3')
        sheet.merge_cells('A4:H4')
        sheet.merge_cells('A5:H5')
        sheet.merge_cells('A6:H6')
        sheet.merge_cells('A7:H7')
        sheet.merge_cells('A8:H8')
        sheet.merge_cells('A9:H9')
        sheet.merge_cells(start_row=len(dados1) + 11, start_column=1, end_row=len(dados1) + 11, end_column=7)
        sheet.merge_cells(start_row=len(dados1) + 12, start_column=1, end_row=len(dados1) + 12, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 14, start_column=1, end_row=total_linhas + 14, end_column=7)
        sheet.merge_cells(start_row=total_linhas + 15, start_column=1, end_row=total_linhas + 15, end_column=7)
        sheet.merge_cells(start_row=total_linhas + 17, start_column=1, end_row=total_linhas + 17, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 18, start_column=1, end_row=total_linhas + 18, end_column=7)

        sheet.merge_cells(start_row=total_linhas + 20, start_column=1, end_row=total_linhas + 20, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 23, start_column=1, end_row=total_linhas + 23, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 24, start_column=1, end_row=total_linhas + 24, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 27, start_column=1, end_row=total_linhas + 27, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 28, start_column=1, end_row=total_linhas + 28, end_column=8)
        

        # Define as células A7 e A9 com fonte negrito
        sheet['A7'].font = fonte_negrito
        sheet['A9'].font = fonte_negrito
        sheet.cell(row=len(dados1) + 11, column=8).font = fonte_negrito
        sheet.cell(row=len(dados1) + 12, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 15, column=1).font = fonte_negrito        
        sheet.cell(row=total_linhas + 20, column=1).font = fonte_padrao
        sheet.cell(row=total_linhas + 23, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 24, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 27, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 28, column=1).font = fonte_negrito


        sheet.cell(row=total_linhas + 14, column=8).style = 'valor_format'
        sheet.cell(row=total_linhas + 15, column=8).style = 'valor_format'
        sheet.cell(row=total_linhas + 18, column=8).style = 'valor_format'
        sheet.cell(row=total_linhas + 14, column=8).font = fonte_negrito
        sheet.cell(row=total_linhas + 15, column=8).font = fonte_negrito
        sheet.cell(row=total_linhas + 17, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 18, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 18, column=8).font = fonte_negrito
        sheet.cell(row=total_linhas + 14, column=8).border = thin_border
        sheet.cell(row=total_linhas + 15, column=8).border = thin_border
        sheet.cell(row=total_linhas + 17, column=1).border = thin_border
        sheet.cell(row=total_linhas + 18, column=8).border = thin_border
        
        
        
        # # Aplica a borda externa às células das colunas 1 a 7 na linha total_linhas + 15
        # for col in range(1, 8):
        #     sheet.cell(row=total_linhas + 14, column=col).border = outer_border

        # # Aplica a borda externa às células das colunas 1 a 7 na linha total_linhas + 15
        # for col in range(1, 9):
        #     sheet.cell(row=total_linhas + 15, column=col).border = outer_border
        # # Aplica a borda externa às células das colunas 1 a 7 na linha total_linhas + 15
        # for col in range(1, 8):
        #     sheet.cell(row=total_linhas + 16, column=col).border = outer_border
        # # Aplica a borda externa às células das colunas 1 a 7 na linha total_linhas + 15
        # for col in range(1, 8):
        #     sheet.cell(row=total_linhas + 15, column=col).border = outer_border        

        # # Aplica a borda externa às células das colunas 1 a 8 na linha total_linhas + 17
        # for col in range(1, 9):
        #     sheet.cell(row=total_linhas + 17, column=col).border = outer_border


                 



        # # tirando borda interna:

        for row in sheet.iter_rows(min_row=1, max_row=6, min_col=1, max_col=8):
            for cell in row:
                if cell.row == 1:
                    cell.border = Border(top=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
                elif cell.row == 6:
                    cell.border = Border(bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
                else:
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'))

        for row in sheet.iter_rows(min_row=self.total_linhas + 19, max_row=self.total_linhas + 28, min_col=1, max_col=8):
            for cell in row:
                cell.border = None

        for row in sheet.iter_rows(min_row=self.total_linhas + 16, max_row=self.total_linhas + 16, min_col=1, max_col=8):
            for cell in row:
                cell.border = Border(bottom=Side(style='thin'), top=Side(style='thin'))        

        
class RelatorioDinheiro(View):
    nome_colunas_excel = {
        'saida': 'Data de Saída',
        'graduacao': 'PST/GRAD',
        'nome_pagante': 'Hóspede',
        'cpf_pagante': 'CPF',
        'qtde_acomp': 'Nr Acomp',
        'uh': 'UH',
        'diarias': 'Dias',
        'valor_total': 'Total (R$)',
    }

    cabecalho_fixo = 9
    cabecalho_dados1 = 1
    cabecalho_dados2 = 1
    linhas_dados1 = None
    linhas_dados2 = None
    mes_texto = None
    total_linhas = None
    total_htm1 = 0
    total_htm2 = 0
    dados1 = None
    dados2 = None
    mes_relatorio = None
    ano_relatorio = None


    def get(self, request, *args, **kwargs):
        self.mes_relatorio = request.GET.get('mes_relatorio')
        self.ano_relatorio = request.GET.get('ano_relatorio')
        forma_pagamento = request.GET.get('forma_pagamento')
        # print("Mês Relatório: auditoria", mes_relatorio)  # Verifique se os valores estão corretos
        # print("Ano Relatório: auditoria", ano_relatorio)
        # print("Forma Pagamento: auditoria", forma_pagamento)

        meses = {
            '1': 'JAN',
            '2': 'FEV',
            '3': 'MAR',
            '4': 'ABR',
            '5': 'MAI',
            '6': 'JUN',
            '7': 'JUL',
            '8': 'AGO',
            '9': 'SET',
            '10': 'OUT',
            '11': 'NOV',
            '12': 'DEZ'
        }

        
        # Verificar se mes_relatorio, ano_relatorio e forma_pagamento não são None
        if self.mes_relatorio is None or self.ano_relatorio is None or forma_pagamento is None:
            return HttpResponseBadRequest("Parâmetros mes_relatorio, ano_relatorio e forma_pagamento são obrigatórios.")

        # Validação dos dados de entrada
        try:
            self.mes_relatorio = int(self.mes_relatorio)
            self.ano_relatorio = int(self.ano_relatorio)
        except ValueError:
            return HttpResponseBadRequest("Mês e ano devem ser números inteiros.")

        if self.mes_relatorio < 1 or self.mes_relatorio > 12:
            return HttpResponseBadRequest("Mês deve estar entre 1 e 12.")

        if self.ano_relatorio < 2024 or self.ano_relatorio > 2030:
            return HttpResponseBadRequest("Ano deve estar entre 2024 e 2030.")
        
        self.mes_texto = meses.get(str(self.mes_relatorio), 'DESC')
        print(self.mes_texto)

        # Realize as consultas usando os parâmetros recebidos
        consultar_htm1 = BaseDados.objects.filter(
            status_reserva="Pago",
            mhex="HTM_01",
            forma_pagamento=forma_pagamento,
            saida__month=self.mes_relatorio,
            saida__year=self.ano_relatorio
        )

        # Consulta para o segundo tipo de dados
        consultar_htm2 = BaseDados.objects.filter(
            status_reserva="Pago",
            mhex="HTM_02",
            forma_pagamento=forma_pagamento,
            saida__month=self.mes_relatorio,
            saida__year=self.ano_relatorio
        )

        if not consultar_htm1.exists() and not consultar_htm2.exists():
            return HttpResponse("Não há lançamentos para o mês solicitado.")
        


        self.total_htm1 = BaseDados.objects.filter(
            status_reserva="Pago",
            mhex="HTM_01",
            forma_pagamento=forma_pagamento,
            saida__month=self.mes_relatorio,
            saida__year=self.ano_relatorio
        ).aggregate(soma=Sum('valor_total'))['soma']  

        self.total_htm2 = BaseDados.objects.filter(
            status_reserva="Pago",
            mhex="HTM_02",
            forma_pagamento=forma_pagamento,
            saida__month=self.mes_relatorio,
            saida__year=self.ano_relatorio
        ).aggregate(soma=Sum('valor_total'))['soma']

        # if not self.total_htm1.exists() and not self.total_htm2.exists():
        #     return HttpResponse("Não há lançamentos para o mês solicitado.")

        # Combine os resultados das consultas em um único conjunto de dados
        self.dados1 = list(consultar_htm1)
        self.dados2 = list(consultar_htm2)

        self.linhas_dados1 = len(self.dados1) 
        self.linhas_dados2 = len(self.dados2)       
        self.total_linhas = self.linhas_dados1 + self.linhas_dados2

        # Gere a planilha
        workbook = self.gerar_planilha(self.dados1, self.dados2, self.linhas_dados1, self.mes_relatorio, self.mes_texto)

        # Configura o Excel
        self.configurar_excel(workbook, self.dados1, self.dados2, self.linhas_dados1, self.linhas_dados2, self.total_linhas, self.ano_relatorio)

        # Configura o Estilo
        self.configurar_estilos(workbook, self.dados1, self.dados2, self.linhas_dados1, self.linhas_dados2, self.total_linhas, self.total_htm1, self.total_htm2, self.ano_relatorio)

        # Retorna a resposta HTTP com a planilha anexada
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="relatorio_{forma_pagamento}.xlsx"'

        # Salva o workbook como uma resposta HTTP
        workbook.save(response)
        return response

    def gerar_planilha(self, dados1, dados2, linhas_dados1, mes_relatorio, mes_texto):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = f'{ self.mes_texto}'

        # INCLUIR CABEÇALHO DADOS 1
        for col, coluna in enumerate(self.nome_colunas_excel.values(), start=1):
            sheet.cell(row=10, column=col).value = coluna

        # INCLUIR DADOS 1
        self.escrever_dados(sheet, 11, self.dados1)

        # INCLUIR CABEÇALHO DADOS 2
        for col, coluna in enumerate(self.nome_colunas_excel.values(), start=1):
            sheet.cell(row=self.linhas_dados1 + 10 + 3, column=col).value = coluna

        # Escreva os dados do segundo conjunto na planilha a partir da linha len(dados1) + 11 + 3
        self.escrever_dados(sheet, self.linhas_dados1 + 10 + 4, self.dados2)

        return workbook

    def escrever_dados(self, sheet, start_row, dados):
        # Escrever os dados
        for row, registro in enumerate(dados, start=start_row):
            for col, coluna in enumerate(self.nome_colunas_excel.keys(), start=1):
                valor = getattr(registro, coluna, None)
                if coluna == 'valor_total' and valor is not None:
                    try:
                        valor = float(valor)
                    except ValueError:
                        valor = None
                sheet.cell(row=row, column=col).value = valor

    def configurar_excel(self, workbook, dados1, dados2, linhas_dados1, linhas_dados2, total_linhas, ano_relatorio):
        # Acessa a planilha ativa
        sheet = workbook.active

        # Larguras das colunas desejadas
        larguras_colunas = {'A': 19, 'B': 13, 'C': 56, 'D': 16.71, 'E': 10, 'F': 9.5, 'G': 9.5, 'H': 17}
        for col, width in larguras_colunas.items():
            sheet.column_dimensions[col].width = width

        # Formatar coluna 'A' (Saida) para dd/mm/aaaa
        date_format = NamedStyle(name='date_format')
        date_format.number_format = 'DD/MM/YYYY'

        # Aplicar estilo à coluna 'A' (Saida)
        for cell in sheet['A']:
            cell.style = date_format

        # Ocultar as linhas de grade
        sheet.sheet_view.showGridLines = False

        # Formatar coluna 'H' (valor_total) para ##.###,##
        valor_format = NamedStyle(name='valor_format')
        valor_format.number_format = '#,##0.00' if '.' in locale.localeconv()['decimal_point'] else '#.##0,00'

        # Aplicar estilo à coluna 'H' (valor_total)
        for cell in sheet['H']:
            cell.style = valor_format

        # Consulta para os dados do banco
        consultar_dados = self.dados1

        # Configura outros estilos, como bordas, fontes, etc.
        self.configurar_estilos(workbook, self.dados1, self.dados2, self.linhas_dados1, self.linhas_dados2, self.total_linhas, self.total_htm1, self.total_htm2, self.ano_relatorio)
        
        # Centralizar as colunas especificadas
        colunas_centralizadas = ['A', 'B', 'D', 'E', 'F', 'G']  # Colunas 'saida', 'graduacao', 'cpf', 'qtde_acomp', 'uh', 'diarias'
        for coluna in colunas_centralizadas:
            for cell in sheet[coluna]:
                cell.alignment = Alignment(horizontal='center')

        # consultar_htm1 = BaseDados.objects.filter(status_reserva="Pago", mhex="HTM_01", forma_pagamento="PIX", saida__month=1)
        # dados1 = list(consultar_htm1)
        # consultar_htm2 = BaseDados.objects.filter(status_reserva="Pago", mhex="HTM_02", forma_pagamento="PIX", saida__month=1)
        # dados2 = list(consultar_htm2)
        # linhas_dados1 = len(dados1) 
        # linhas_dados2 = len(dados2)       
         
        sheet.cell(row=self.total_linhas + 15, column=1).alignment = Alignment(horizontal='left', vertical='center')
        sheet.cell(row=self.total_linhas + 17, column=1).alignment = Alignment(horizontal='left', vertical='center')
        sheet.cell(row=self.total_linhas + 18, column=1).alignment = Alignment(horizontal='left', vertical='center')        

    # def escrever_dados(self, sheet, start_row, dados):
    #     # Defina as colunas desejadas diretamente aqui
    #     colunas_desejadas = ['saida', 'graduacao', 'nome', 'cpf', 'qtde_acomp', 'uh', 'diarias', 'valor_total']

    #     # Escrever os dados
    #     for row, registro in enumerate(dados, start=start_row):
    #         for col, coluna in enumerate(colunas_desejadas, start=1):
    #             valor = getattr(registro, coluna, None)
    #             if coluna == 'valor_total' and valor is not None:
    #                 try:
    #                     valor = float(valor)
    #                 except ValueError:
    #                     valor = None
    #             sheet.cell(row=row, column=col).value = valor

    def configurar_estilos(self, workbook, dados1, dados2, linhas_dados1, linhas_dados2, total_linhas, total_htm1, total_htm2, ano_relatorio):
        sheet = workbook.active
        # consultar_htm1 = BaseDados.objects.filter(status_reserva="Pago", mhex="HTM_01", forma_pagamento="PIX", saida__month=1)
        # dados1 = list(consultar_htm1)
        # consultar_htm2 = BaseDados.objects.filter(status_reserva="Pago", mhex="HTM_02", forma_pagamento="PIX", saida__month=1)
        # dados2 = list(consultar_htm2)
        
        # linhas_dados1 = len(dados1) 
        # linhas_dados2 = len(dados2)       
        # total_linhas = linhas_dados1 + linhas_dados2

        
        # total_linhas = len(dados1) + len(dados2)
        sheet.cell(row=self.total_linhas + 15, column=1).alignment = Alignment(horizontal='left', vertical='center')
        
        meses2 = {
            'JAN': 'janeiro',
            'FEV': 'fevereiro',
            'MAR': 'março',
            'ABR': 'abril',
            'MAI': 'maio',
            'JUN': 'junho',
            'JUL': 'julho',
            'AGO': 'agosto',
            'SET': 'setembro',
            'OUT': 'outubro',
            'NOV': 'novembro',
            'DEZ': 'dezembro'
        }

        self.mes2_texto = meses2.get(self.mes_texto)

        # Defina os estilos de borda
        dotted_border = Border(left=Side(style='dotted'),
                            right=Side(style='dotted'),
                            top=Side(style='dotted'),
                            bottom=Side(style='dotted'))

        thin_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))

        # Define as fontes
        fonte_padrao = Font(name='Times New Roman', size=11, bold=False)
        fonte_negrito = Font(name='Times New Roman', size=11, bold=True)

        # Define o preenchimento cinza claro para o cabeçalho
        gray_fill = PatternFill(start_color='00CCCCCC',
                                end_color='00CCCCCC',
                                fill_type='solid')

        # Aplica os estilos para todas as células
        for row in sheet.iter_rows():
            for cell in row:
                # Aplica a borda
                cell.border = thin_border

                # Aplica a fonte padrão
                cell.font = fonte_padrao

        # Aplica estilos para o cabeçalho
        for row in sheet.iter_rows(min_row=10, max_row=10):
            for cell in row:
                cell.font = fonte_negrito
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = gray_fill

        # Aplica estilos para o cabeçalho 2
        for row in sheet.iter_rows(min_row=len(dados1) + 11 + 2, max_row=len(dados1) + 11 + 2):
            for cell in row:
                cell.font = fonte_negrito
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = gray_fill

        # # Defina o estilo de borda externa
        # outer_border = Border(left=Side(style='thin'), right=Side(style='thin'),
        #                     top=Side(style='thin'), bottom=Side(style='thin'))
        # # Aplica a borda externa às células das colunas 1 a 8 na linha total_linhas + 18
        # for col in range(1, 9):
        #     sheet.cell(row=self.total_linhas + 13, column=col).border = outer_border
        #     sheet.cell(row=self.total_linhas + 14, column=col).border = outer_border 
        #     sheet.cell(row=self.total_linhas + 15, column=col).border = outer_border 
        #     sheet.cell(row=self.total_linhas + 16, column=col).border = outer_border 
        #     sheet.cell(row=self.total_linhas + 17, column=col).border = outer_border 
        #     sheet.cell(row=self.total_linhas + 18, column=col).border = outer_border 
            
        # Seu texto
        texto = """MINISTERIO DA DEFESA
EXÉRCITO BRASILEIRO
COMANDO MILITAR DO OESTE
4ª BRIGADA DE CAVALARIA MECANIZADA
11º REGIMENTO DE CAVALARIA MECANIZADO
REGIMENTO MARECHAL DUTRA
"""

        # Insere o texto na célula A1
        sheet['A1'] = """MINISTERIO DA DEFESA"""
        sheet['A2'] = """EXÉRCITO BRASILEIRO"""   
        sheet['A3'] = """COMANDO MILITAR DO OESTE"""   
        sheet['A4'] = """4ª BRIGADA DE CAVALARIA MECANIZADA"""   
        sheet['A5'] = """11º REGIMENTO DE CAVALARIA MECANIZADO"""   
        sheet['A6'] = """REGIMENTO MARECHAL DUTRA"""   
        sheet['A7'] = """RELATÓRIO DE AUDITORIA DOS MEIOS DE HOSPEDAGEM DO EXÉRCITO NA GUARNIÇÃO DE PONTA PORÃ - MS"""   
        sheet['A8'] = f"Trata o presente relatório sobre auditoria realizada nos meios de Hospedagem do Exército da Guarnição de Ponta Porã relativo ao mês de {self.mes2_texto} de {self.ano_relatorio}."  
        sheet['A9'] = """HTM 1"""
        sheet.cell(row=self.linhas_dados1 + 11, column=8).value = total_htm1
        sheet.cell(row=self.linhas_dados1 + 12, column=1).value = "HTM 2"
        
        
        
        
       
        sheet.cell(row=total_linhas + 14, column=8).value = total_htm2
        
        total_geral = total_htm1 + total_htm2
        sheet.cell(row=total_linhas + 15, column=1).value = "TOTAL EM DINHEIRO"
        sheet.cell(row=total_linhas + 15, column=8).value = total_geral

        sheet.cell(row=total_linhas + 17, column=1).value = "2. RECOLHIDO AO FUNDO DO EXÉRCITO"
        sheet.cell(row=total_linhas + 18, column=1).value = "DEP. DINHEIRO"
        sheet.cell(row=total_linhas + 18, column=8).value = total_geral

        # Define o idioma como português brasileiro
        locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
        # Obter a data de hoje
        data_hoje = datetime.now()
        # Formatar a data como "dia de mês de ano"
        data_formatada = data_hoje.strftime("%d de %B de %Y")

        # Inserir a data formatada na célula desejada
        sheet.cell(row=total_linhas + 20, column=1).value = f"Ponta Porã-MS, {data_formatada}"
        sheet.cell(row=total_linhas + 23, column=1).value = "EDER SCHWEIGERT FONSECA - Cap"
        sheet.cell(row=total_linhas + 24, column=1).value = "Gestor HT 11º R C Mec"
        sheet.cell(row=total_linhas + 27, column=1).value = "CHRYSTIAN HENRY BRITO CARDOSO - Maj"
        sheet.cell(row=total_linhas + 28, column=1).value = "Fiscal Administrativo 11º R C Mec"

        

        

        # Ajusta o alinhamento para envolver o texto e centralizá-lo
        sheet['A1'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A2'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A3'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A4'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A5'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A6'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A7'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A8'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
        sheet['A9'].alignment = Alignment(wrapText=False, horizontal='center', vertical='center')
        sheet.cell(row=len(dados1) + 12, column=1).alignment = Alignment(wrapText=False, horizontal='center', vertical='center')

        

        # Mescla as células A1:H1
        sheet.merge_cells('A1:H1')
        sheet.merge_cells('A2:H2')
        sheet.merge_cells('A3:H3')
        sheet.merge_cells('A4:H4')
        sheet.merge_cells('A5:H5')
        sheet.merge_cells('A6:H6')
        sheet.merge_cells('A7:H7')
        sheet.merge_cells('A8:H8')
        sheet.merge_cells('A9:H9')
        sheet.merge_cells(start_row=len(dados1) + 11, start_column=1, end_row=len(dados1) + 11, end_column=7)
        sheet.merge_cells(start_row=len(dados1) + 12, start_column=1, end_row=len(dados1) + 12, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 14, start_column=1, end_row=total_linhas + 14, end_column=7)
        sheet.merge_cells(start_row=total_linhas + 15, start_column=1, end_row=total_linhas + 15, end_column=7)
        sheet.merge_cells(start_row=total_linhas + 17, start_column=1, end_row=total_linhas + 17, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 18, start_column=1, end_row=total_linhas + 18, end_column=7)

        sheet.merge_cells(start_row=total_linhas + 20, start_column=1, end_row=total_linhas + 20, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 23, start_column=1, end_row=total_linhas + 23, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 24, start_column=1, end_row=total_linhas + 24, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 27, start_column=1, end_row=total_linhas + 27, end_column=8)
        sheet.merge_cells(start_row=total_linhas + 28, start_column=1, end_row=total_linhas + 28, end_column=8)
        

        # Define as células A7 e A9 com fonte negrito
        sheet['A7'].font = fonte_negrito
        sheet['A9'].font = fonte_negrito
        sheet.cell(row=len(dados1) + 11, column=8).font = fonte_negrito
        sheet.cell(row=len(dados1) + 12, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 15, column=1).font = fonte_negrito        
        sheet.cell(row=total_linhas + 20, column=1).font = fonte_padrao
        sheet.cell(row=total_linhas + 23, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 24, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 27, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 28, column=1).font = fonte_negrito


        sheet.cell(row=total_linhas + 14, column=8).style = 'valor_format'
        sheet.cell(row=total_linhas + 15, column=8).style = 'valor_format'
        sheet.cell(row=total_linhas + 18, column=8).style = 'valor_format'
        sheet.cell(row=total_linhas + 14, column=8).font = fonte_negrito
        sheet.cell(row=total_linhas + 15, column=8).font = fonte_negrito
        sheet.cell(row=total_linhas + 17, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 18, column=1).font = fonte_negrito
        sheet.cell(row=total_linhas + 18, column=8).font = fonte_negrito
        sheet.cell(row=total_linhas + 14, column=8).border = thin_border
        sheet.cell(row=total_linhas + 15, column=8).border = thin_border
        sheet.cell(row=total_linhas + 17, column=1).border = thin_border
        sheet.cell(row=total_linhas + 18, column=8).border = thin_border
        
        
        
        # # Aplica a borda externa às células das colunas 1 a 7 na linha total_linhas + 15
        # for col in range(1, 8):
        #     sheet.cell(row=total_linhas + 14, column=col).border = outer_border

        # # Aplica a borda externa às células das colunas 1 a 7 na linha total_linhas + 15
        # for col in range(1, 9):
        #     sheet.cell(row=total_linhas + 15, column=col).border = outer_border
        # # Aplica a borda externa às células das colunas 1 a 7 na linha total_linhas + 15
        # for col in range(1, 8):
        #     sheet.cell(row=total_linhas + 16, column=col).border = outer_border
        # # Aplica a borda externa às células das colunas 1 a 7 na linha total_linhas + 15
        # for col in range(1, 8):
        #     sheet.cell(row=total_linhas + 15, column=col).border = outer_border        

        # # Aplica a borda externa às células das colunas 1 a 8 na linha total_linhas + 17
        # for col in range(1, 9):
        #     sheet.cell(row=total_linhas + 17, column=col).border = outer_border


                 



        # # tirando borda interna:

        for row in sheet.iter_rows(min_row=1, max_row=6, min_col=1, max_col=8):
            for cell in row:
                if cell.row == 1:
                    cell.border = Border(top=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
                elif cell.row == 6:
                    cell.border = Border(bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
                else:
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'))

        for row in sheet.iter_rows(min_row=self.total_linhas + 19, max_row=self.total_linhas + 28, min_col=1, max_col=8):
            for cell in row:
                cell.border = None

        for row in sheet.iter_rows(min_row=self.total_linhas + 16, max_row=self.total_linhas + 16, min_col=1, max_col=8):
            for cell in row:
                cell.border = Border(bottom=Side(style='thin'), top=Side(style='thin'))        
